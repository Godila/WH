from io import BytesIO
from typing import Optional
import importlib

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.product import Product
from app.models.stock import Stock
from app.models.defect_stock import DefectStock

_schemas = importlib.import_module("app.schemas.import_schema")
ExcelImportRow = _schemas.ExcelImportRow
ExcelImportError = _schemas.ExcelImportError
ExcelImportResult = _schemas.ExcelImportResult


class ExcelImportService:
    COLUMN_MAPPING = {
        "Баркод": "barcode",
        "Артикул продавца": "seller_sku",
        "Размер": "size",
        "Бренд": "brand",
        "АКТУАЛЬНЫЙ ОСТАТОК": "stock_quantity",
        "БРАКИ": "defect_quantity",
    }
    BATCH_SIZE = 500

    def __init__(self, db: AsyncSession):
        self.db = db

    @staticmethod
    def _generate_gtin(barcode: str) -> str:
        if len(barcode) in (13, 14) and barcode.isdigit():
            return barcode.zfill(14)
        padded = barcode[:11].ljust(11, "0")
        return f"IMP{padded}"

    async def parse_excel(
        self, file_content: bytes
    ) -> tuple[list[ExcelImportRow], list[ExcelImportError]]:
        rows: list[ExcelImportRow] = []
        errors: list[ExcelImportError] = []
        
        try:
            workbook = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
        except Exception as e:
            errors.append(ExcelImportError(
                row_number=0,
                field="file",
                message=f"Failed to open Excel file: {str(e)}"
            ))
            return rows, errors

        sheet = None
        for ws in workbook.worksheets:
            if ws.title == "Сводная":
                sheet = ws
                break
        
        if sheet is None:
            errors.append(ExcelImportError(
                row_number=0,
                field="sheet",
                message="Sheet 'Сводная' not found"
            ))
            workbook.close()
            return rows, errors

        header_row = next(sheet.iter_rows(min_row=1, max_row=1), None)
        if header_row is None:
            errors.append(ExcelImportError(
                row_number=1,
                field="header",
                message="Empty file - no header row"
            ))
            workbook.close()
            return rows, errors

        column_indices: dict[str, int] = {}
        for idx, cell in enumerate(header_row):
            header_value = str(cell.value).strip() if cell.value else ""
            if header_value in self.COLUMN_MAPPING:
                field_name = self.COLUMN_MAPPING[header_value]
                column_indices[field_name] = idx

        if "barcode" not in column_indices:
            errors.append(ExcelImportError(
                row_number=1,
                field="barcode",
                message="Required column 'Баркод' not found"
            ))
            workbook.close()
            return rows, errors

        for row_num, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            if all(cell.value is None for cell in row):
                continue

            def get_cell_value(field: str) -> Optional[str]:
                if field not in column_indices:
                    return None
                idx = column_indices[field]
                if idx >= len(row):
                    return None
                val = row[idx].value
                return str(val).strip() if val is not None else None

            def get_int_value(field: str) -> int:
                if field not in column_indices:
                    return 0
                idx = column_indices[field]
                if idx >= len(row):
                    return 0
                val = row[idx].value
                if val is None:
                    return 0
                try:
                    return int(float(val))
                except (ValueError, TypeError):
                    return 0

            barcode = get_cell_value("barcode")
            if not barcode:
                errors.append(ExcelImportError(
                    row_number=row_num,
                    field="barcode",
                    message="Barcode is required"
                ))
                continue

            rows.append(ExcelImportRow(
                barcode=barcode,
                seller_sku=get_cell_value("seller_sku"),
                size=get_cell_value("size"),
                brand=get_cell_value("brand"),
                stock_quantity=get_int_value("stock_quantity"),
                defect_quantity=get_int_value("defect_quantity"),
                row_number=row_num,
            ))

        workbook.close()
        return rows, errors

    async def validate_rows(self, rows: list[ExcelImportRow]) -> list[ExcelImportError]:
        errors: list[ExcelImportError] = []
        
        seen_barcodes: dict[str, int] = {}
        for row in rows:
            if row.barcode in seen_barcodes:
                errors.append(ExcelImportError(
                    row_number=row.row_number,
                    field="barcode",
                    message=f"Duplicate barcode, first occurrence at row {seen_barcodes[row.barcode]}"
                ))
            else:
                seen_barcodes[row.barcode] = row.row_number

            if row.stock_quantity < 0:
                errors.append(ExcelImportError(
                    row_number=row.row_number,
                    field="stock_quantity",
                    message="Stock quantity cannot be negative"
                ))

            if row.defect_quantity < 0:
                errors.append(ExcelImportError(
                    row_number=row.row_number,
                    field="defect_quantity",
                    message="Defect quantity cannot be negative"
                ))

        return errors

    async def import_batch(self, rows: list[ExcelImportRow]) -> tuple[int, int]:
        created = 0
        updated = 0

        for row in rows:
            result = await self.db.execute(
                select(Product).where(Product.barcode == row.barcode)
            )
            product = result.scalar_one_or_none()

            if product:
                product.seller_sku = row.seller_sku
                product.size = row.size
                product.brand = row.brand
                product.is_deleted = False

                product.stock.quantity = row.stock_quantity
                product.defect_stock.quantity = row.defect_quantity
                updated += 1
            else:
                gtin = self._generate_gtin(row.barcode)
                product = Product(
                    barcode=row.barcode,
                    gtin=gtin,
                    seller_sku=row.seller_sku,
                    size=row.size,
                    brand=row.brand,
                )
                self.db.add(product)
                await self.db.flush()

                stock = Stock(
                    product_id=product.id,
                    quantity=row.stock_quantity,
                )
                self.db.add(stock)

                defect_stock = DefectStock(
                    product_id=product.id,
                    quantity=row.defect_quantity,
                )
                self.db.add(defect_stock)
                created += 1

        await self.db.commit()
        return created, updated

    async def import_from_file(self, file_content: bytes) -> ExcelImportResult:
        rows, parse_errors = await self.parse_excel(file_content)
        if parse_errors:
            return ExcelImportResult(
                total_rows=0,
                created=0,
                updated=0,
                errors=parse_errors,
                success=False,
            )

        if not rows:
            return ExcelImportResult(
                total_rows=0,
                created=0,
                updated=0,
                errors=[],
                success=True,
            )

        validation_errors = await self.validate_rows(rows)
        if validation_errors:
            return ExcelImportResult(
                total_rows=len(rows),
                created=0,
                updated=0,
                errors=validation_errors,
                success=False,
            )

        total_created = 0
        total_updated = 0

        for i in range(0, len(rows), self.BATCH_SIZE):
            batch = rows[i : i + self.BATCH_SIZE]
            created, updated = await self.import_batch(batch)
            total_created += created
            total_updated += updated

        return ExcelImportResult(
            total_rows=len(rows),
            created=total_created,
            updated=total_updated,
            errors=[],
            success=True,
        )
